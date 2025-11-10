import os
import pandas as pd
import datetime
from openpyxl import Workbook, load_workbook
import calendar
import requests
from openpyxl.styles import PatternFill

TAGES_SOLL_STUNDEN = 8
PAUSE_IN_STUNDEN = 0.75

weekday_names = ['Montag', 'Dienstag', 'Mittwoch', 'Donnerstag', 'Freitag', 'Samstag', 'Sonntag']
feiertage_filter = ['08-08', '08-15', '11-19']

year = datetime.datetime.now().year
# URL der API
url = f"https://feiertage-api.de/api/?jahr={year}&nur_land=BY"

# API-Aufruf
response = requests.get(url)

# Antwort in JSON umwandeln
data = response.json()

# Leere Liste für die Feiertage
feiertage = []

holidays = response.json()

# Iteriere über alle Feiertage
for holiday, date in holidays.items():
    # Füge das Datum des Feiertags der Liste hinzu
    if date["datum"][5:] not in feiertage_filter:
        feiertage.append(date["datum"])

CALENDAR_FILE = "calendar.xlsx"
existing_wb = load_workbook(CALENDAR_FILE) if os.path.exists(CALENDAR_FILE) else None

# Erstelle ein leeres Workbook
wb = Workbook()

# Benenne das Standard-Tab "Sheet" in "Übersicht" um
if "Sheet" in wb.sheetnames:
    wb["Sheet"].title = "Übersicht"

# Füge die Summen der Monate in die Übersicht ein
uebersicht_ws = wb["Übersicht"]

# Schreibe die Kopfzeile in die Übersicht (inklusive der neuen Spalten für Abwesenheiten)
uebersicht_ws.append(["Monat", "Summe Soll", "Summe Arbeitszeit", "Summe Überstunden", "Urlaub", "Krank", "Dienstreise", "Gleittage"])

# Variable zur Speicherung der Überstunden des Vormonats
ueberstunden_vormonat = 0

# Definiere die Farben
grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Grau für Wochenende
pink_fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")  # Rosa für Feiertage
turquoise_fill = PatternFill(start_color="40E0D0", end_color="40E0D0", fill_type="solid")  # Türkis für Urlaub

# Iteriere über alle Monate des aktuellen Jahres
for month in range(1, 13):
    # Erstelle einen leeren DataFrame mit den Spalten "Datum", "Gekommen", "Gehzeit", "Pause", "Arbeitszeit", "Soll", "Überstunden"
    df = pd.DataFrame(columns=['Datum', 'Tag', 'Gekommen', 'Gehzeit', 'Pause', 'Arbeitszeit', 'Soll', 'Überstunden'])

    # Bestimme die Anzahl der Tage im aktuellen Monat
    num_days = calendar.monthrange(datetime.datetime.now().year, month)[1]

    # Lese vorhandene Einträge aus bestehender Datei, um manuelle Zeiten zu erhalten
    existing_rows = {}
    if existing_wb and calendar.month_name[month] in existing_wb.sheetnames:
        existing_month_ws = existing_wb[calendar.month_name[month]]
        for row_idx in range(2, existing_month_ws.max_row):
            day_cell = existing_month_ws.cell(row=row_idx, column=1).value
            if day_cell is None:
                continue
            if isinstance(day_cell, datetime.date):
                day_key = day_cell.day
            else:
                try:
                    day_key = int(day_cell)
                except (TypeError, ValueError):
                    continue
            existing_rows[day_key] = {
                'Gekommen': existing_month_ws.cell(row=row_idx, column=3).value,
                'Gehzeit': existing_month_ws.cell(row=row_idx, column=4).value,
                'Pause': existing_month_ws.cell(row=row_idx, column=5).value,
                'Arbeitszeit': existing_month_ws.cell(row=row_idx, column=6).value,
                'Soll': existing_month_ws.cell(row=row_idx, column=7).value,
                'Überstunden': existing_month_ws.cell(row=row_idx, column=8).value
            }

    # Iteriere über alle Tage des aktuellen Monats
    for day in pd.date_range(start=f'{datetime.datetime.now().year}-{month:02d}-01', end=f'{datetime.datetime.now().year}-{month:02d}-{num_days}'):
        day_is_holiday = False
        for feiertag in feiertage:
            if feiertag == day._date_repr:
                day_is_holiday = True
                break

        # Überprüfe, ob der aktuelle Tag ein Wochentag ist
        date = day.date()
        if day_is_holiday:
            # Feiertag: Setze alle Werte auf 0
            row_data = {
                'Tag': weekday_names[day.dayofweek],
                'Datum': date.day,
                'Gekommen': 0,
                'Gehzeit': 0,
                'Pause': 0,
                'Arbeitszeit': 0,
                'Soll': 0,
                'Überstunden': 0
            }
        elif day.weekday() in [5, 6]:  # Samstag oder Sonntag
            # Wochenende: Setze alle Werte auf 0
            row_data = {
                'Tag': weekday_names[day.dayofweek],
                'Datum': date.day,
                'Gekommen': 0,
                'Gehzeit': 0,
                'Pause': 0,
                'Arbeitszeit': 0,
                'Soll': 0,
                'Überstunden': 0
            }
        else:
            # Setze die Gekommen-Zeit auf 7:30
            time = datetime.datetime.combine(date, datetime.time(7, 30))

            # Setze die Gehzeit auf 16:30
            gehzeit = datetime.datetime.combine(date, datetime.time(16, 30))

            # Berechne die Arbeitszeit zwischen der Uhrzeit und der Gehzeit in Minuten
            diff = (gehzeit - time).total_seconds() / 60

            # Wandle die Arbeitszeit in Stunden um und ziehe die Pause ab
            diff_hours = diff / 60
            arbeitszeit = diff_hours - PAUSE_IN_STUNDEN
            ueberstunden = arbeitszeit - TAGES_SOLL_STUNDEN * (len(df) + 1)  # Berechne Überstunden

            # Erstelle eine neue Zeile als dict
            row_data = {
                'Tag': weekday_names[day.dayofweek],
                'Datum': date.day,
                'Gekommen': '7:30',
                'Gehzeit': '16:30',
                'Pause': PAUSE_IN_STUNDEN,
                'Arbeitszeit': arbeitszeit,
                'Soll': TAGES_SOLL_STUNDEN,
                'Überstunden': ueberstunden
            }

        existing_row = existing_rows.get(date.day)
        if existing_row:
            for key, value in existing_row.items():
                if value is not None:
                    row_data[key] = value

        # Füge die neue Zeile ohne concat hinzu, um FutureWarning zu vermeiden
        df.loc[len(df)] = row_data

    ueberstunden_vormonat = ueberstunden + ueberstunden_vormonat
    # Erstelle ein neues Arbeitsblatt für den aktuellen Monat
    ws = wb.create_sheet(title=calendar.month_name[month])

    # Schreibe die Spaltenüberschriften manuell in das Arbeitsblatt
    for col, value in enumerate(df.columns):
        ws.cell(row=1, column=col + 1).value = value

    # Zusatzspalten für Hinweise, Kranktage und Dienstreisen
    extra_col_index = len(df.columns)
    ws.cell(row=1, column=extra_col_index + 1).value = "Hinweis"
    ws.cell(row=1, column=extra_col_index + 2).value = "Urlaubstage"
    ws.cell(row=1, column=extra_col_index + 3).value = "Gleittage"
    ws.cell(row=1, column=extra_col_index + 4).value = "Kranktage"
    ws.cell(row=1, column=extra_col_index + 5).value = "Dienstreisen"

    # Schreibe den DataFrame manuell in das Arbeitsblatt
    for row_idx, row in enumerate(df.iterrows(), start=2):
        for col_idx, value in enumerate(row[1]):
            cell = ws.cell(row=row_idx, column=col_idx + 1)
            cell.value = value

            # Färbe die Zeile entsprechend
            if row[1]['Tag'] in ['Samstag', 'Sonntag']:
                cell.fill = grey_fill  # Grau für Wochenende
            elif f"{datetime.datetime.now().year}-{month:02d}-{row[1]['Datum']:02d}" in feiertage:
                cell.fill = pink_fill  # Rosa für Feiertage

    # Füge Formeln für Arbeitszeit, Summe Soll, Summe Ist und Überstunden hinzu
    for row_idx in range(2, len(df) + 2):
        ws[f'F{row_idx}'] = f'=IF(OR(C{row_idx}="U", C{row_idx}="K", C{row_idx}="D", D{row_idx}=""), {TAGES_SOLL_STUNDEN}, IF(C{row_idx}="G", 0, (D{row_idx}-C{row_idx})*24 - E{row_idx}))' # Arbeitszeit inkl. Sonderfälle
        ws[f'H{row_idx}'] = f'=IF(OR(C{row_idx}="U", C{row_idx}="K", C{row_idx}="D"), 0, F{row_idx} - G{row_idx})' # Überstunden ohne Abwesenheiten
        if row_idx <= 2:
            if month != 1:
                prev_month_sheet = wb[calendar.month_name[month - 1]]
                last_row_prev_month = len(prev_month_sheet['A'])
                month_name = calendar.month_name[month - 1]

    # Füge Summen für Soll, Arbeitszeit und Überstunden in die letzte Zeile ein
    last_row = len(df) + 2
    ws[f'F{last_row}'] = f'=SUM(F2:F{last_row - 1})' # Summe Arbeitszeit
    ws[f'G{last_row}'] = f'=SUM(G2:G{last_row - 1})' # Summe Soll
    ws[f'H{last_row}'] = f'=SUM(H2:H{last_row - 1})' # Summe Überstunden

    # Entferne alte Summenanzeigen für Urlaub und Gleittage in den Spalten C und D
    ws[f'C{last_row}'] = None
    ws[f'D{last_row}'] = None
    # Hinweistext zu den Kürzeln in Spalte I
    ws[f'I{last_row}'] = "Hinweis: U=Urlaub (J), G=Gleittag (K), K=Krank (L), D=Dienstreise (M)"
    # Berechne die Anzahl der Urlaubstage (Zellen in Spalte "Gekommen" mit "U")
    ws[f'J{last_row}'] = f'=COUNTIF(C2:C{last_row - 1}, "U")'
    # Berechne die Anzahl der Gleittage (Zellen in Spalte "Gekommen" mit "G")
    ws[f'K{last_row}'] = f'=COUNTIF(C2:C{last_row - 1}, "G")'
    # Anzahl der Krankheitstage (Spalte "Gekommen" mit "K")
    ws[f'L{last_row}'] = f'=COUNTIF(C2:C{last_row - 1}, "K")'
    # Anzahl der Dienstreisen (Spalte "Gekommen" mit "D")
    ws[f'M{last_row}'] = f'=COUNTIF(C2:C{last_row - 1}, "D")'
    # Optional: Beschriftung für die Summenzeile
    ws[f'B{last_row}'] = None
    # Optional: Beschriftung für die Summenzeile
    ws[f'E{last_row}'] = "Summen"


# Iteriere über alle Monats-Tabs und berechne die Summen
for month in range(1, 13):
    month_name = calendar.month_name[month]
    if month_name in wb.sheetnames:
        month_ws = wb[month_name]
        last_row = len(month_ws['A'])  # Letzte Zeile des Monats-Tabs

        # Füge die Summen und die Urlaubstage in die Übersicht ein
        uebersicht_ws.append([
            month_name,
            f'={month_name}!G{last_row}',  # Summe Soll
            f'={month_name}!F{last_row}',  # Summe Arbeitszeit
            f'={month_name}!H{last_row}',  # Summe Überstunden
            f'={month_name}!J{last_row}',  # Anzahl der Urlaubstage
            f'={month_name}!L{last_row}',  # Anzahl der Kranktage
            f'={month_name}!M{last_row}',  # Anzahl der Dienstreisen
            f'={month_name}!K{last_row}'  # Anzahl der Gleittage
        ])

# Füge die Summenzeile in die Übersicht ein
last_row_uebersicht = len(uebersicht_ws['A']) + 1  # Nächste freie Zeile in der Übersicht
uebersicht_ws[f'A{last_row_uebersicht}'] = "Summen"
uebersicht_ws[f'B{last_row_uebersicht}'] = f'=SUM(B2:B{last_row_uebersicht - 1})'  # Summe Soll
uebersicht_ws[f'C{last_row_uebersicht}'] = f'=SUM(C2:C{last_row_uebersicht - 1})'  # Summe Arbeitszeit
uebersicht_ws[f'D{last_row_uebersicht}'] = f'=SUM(D2:D{last_row_uebersicht - 1})'  # Summe Überstunden
uebersicht_ws[f'E{last_row_uebersicht}'] = f'=SUM(E2:E{last_row_uebersicht - 1})'  # Summe Urlaubstage
uebersicht_ws[f'F{last_row_uebersicht}'] = f'=SUM(F2:F{last_row_uebersicht - 1})'  # Summe Kranktage
uebersicht_ws[f'G{last_row_uebersicht}'] = f'=SUM(G2:G{last_row_uebersicht - 1})'  # Summe Dienstreisen
uebersicht_ws[f'H{last_row_uebersicht}'] = f'=SUM(H2:H{last_row_uebersicht - 1})'  # Summe Gleittage

# Speichere das Workbook
wb.save(CALENDAR_FILE)