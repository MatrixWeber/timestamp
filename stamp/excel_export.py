"""Excel-Export — generiert Zeiterfassungs-Excel im bisherigen Format."""

import calendar
from datetime import date, time
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

from stamp.db import get_session, Stamp, get_config_float, get_config_int, get_config
from stamp.holidays import get_holidays_for_year
from stamp.service import get_year, get_vacation_balance, get_overtime_total

WEEKDAYS_DE = ["Montag", "Dienstag", "Mittwoch", "Donnerstag", "Freitag", "Samstag", "Sonntag"]
MONTHS_DE = [
    "", "Januar", "Februar", "März", "April", "Mai", "Juni",
    "Juli", "August", "September", "Oktober", "November", "Dezember",
]

# Farben
GREY_FILL = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
PINK_FILL = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
TURQUOISE_FILL = PatternFill(start_color="40E0D0", end_color="40E0D0", fill_type="solid")
KRANK_FILL = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")
GLEITTAG_FILL = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
DIENSTREISE_FILL = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
BOLD_FONT = Font(bold=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

TYPE_MAP = {
    "VACATION": "U",
    "SICK": "K",
    "FLEX": "G",
    "TRAVEL": "D",
}


def _style_header(ws, row, max_col):
    """Formatiert die Kopfzeile."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER


def _time_to_excel(t: time | None) -> float | str | None:
    """Konvertiert time zu Excel-Zeitformat (Bruch eines Tages)."""
    if t is None:
        return None
    return (t.hour * 3600 + t.minute * 60 + t.second) / 86400


def export_excel(year: int | None = None, output_path: str | None = None) -> Path:
    """Exportiert die Zeiterfassung als Excel-Datei."""
    if year is None:
        year = date.today().year

    if output_path is None:
        output_path = f"Stundenschreibung{year}.xlsx"
    output = Path(output_path)

    target_hours = get_config_float("target_hours")
    default_pause = get_config_float("default_pause")
    holidays = get_holidays_for_year(year)

    # Alle Einträge laden
    entries = get_year(year)
    entry_map = {e.date: e for e in entries}

    wb = Workbook()

    # Standard-Sheet umbenennen
    if "Sheet" in wb.sheetnames:
        wb["Sheet"].title = "Übersicht"
    uebersicht = wb["Übersicht"]

    # Übersicht-Header
    overview_headers = ["Monat", "Summe Soll", "Summe Arbeitszeit", "Summe Überstunden",
                        "Urlaub", "Krank", "Dienstreise", "Gleittage"]
    for col, header in enumerate(overview_headers, 1):
        uebersicht.cell(row=1, column=col, value=header)
    _style_header(uebersicht, 1, len(overview_headers))

    # Monats-Sheets erstellen
    for month in range(1, 13):
        month_name = MONTHS_DE[month]
        ws = wb.create_sheet(title=month_name)

        # Header
        headers = ["Datum", "Tag", "Gekommen", "Gehzeit", "Pause", "Arbeitszeit",
                    "Soll", "Überstunden", "Hinweis", "Urlaubstage", "Gleittage",
                    "Kranktage", "Dienstreisen"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        _style_header(ws, 1, len(headers))

        num_days = calendar.monthrange(year, month)[1]
        row_idx = 2

        for day_num in range(1, num_days + 1):
            day = date(year, month, day_num)
            wd = WEEKDAYS_DE[day.weekday()]
            is_weekend = day.weekday() >= 5
            is_holiday = day in holidays

            entry = entry_map.get(day)

            ws.cell(row=row_idx, column=1, value=day_num)
            ws.cell(row=row_idx, column=2, value=wd)

            if is_weekend:
                # Wochenende: alles 0, grau
                for col in range(3, 9):
                    ws.cell(row=row_idx, column=col, value=0)
                for col in range(1, 14):
                    ws.cell(row=row_idx, column=col).fill = GREY_FILL
                    ws.cell(row=row_idx, column=col).border = THIN_BORDER

            elif is_holiday:
                # Feiertag: alles 0, pink
                for col in range(3, 9):
                    ws.cell(row=row_idx, column=col, value=0)
                ws.cell(row=row_idx, column=9, value=holidays[day])
                for col in range(1, 14):
                    ws.cell(row=row_idx, column=col).fill = PINK_FILL
                    ws.cell(row=row_idx, column=col).border = THIN_BORDER

            elif entry and entry.type != "WORK":
                # Abwesenheit
                kuerzel = TYPE_MAP.get(entry.type, "")
                ws.cell(row=row_idx, column=3, value=kuerzel)
                ws.cell(row=row_idx, column=4, value="")
                ws.cell(row=row_idx, column=5, value=0)
                ws.cell(row=row_idx, column=6, value=target_hours)
                ws.cell(row=row_idx, column=7, value=target_hours)
                ws.cell(row=row_idx, column=8, value=0)
                for col in range(1, 14):
                    ws.cell(row=row_idx, column=col).border = THIN_BORDER

            elif entry and entry.stamp_in:
                # Arbeitstag mit Stempel
                in_cell = ws.cell(row=row_idx, column=3)
                in_cell.value = _time_to_excel(entry.stamp_in)
                in_cell.number_format = "HH:MM"

                if entry.stamp_out:
                    out_cell = ws.cell(row=row_idx, column=4)
                    out_cell.value = _time_to_excel(entry.stamp_out)
                    out_cell.number_format = "HH:MM"

                ws.cell(row=row_idx, column=5, value=entry.pause)
                # Formeln für Arbeitszeit und Überstunden
                ws.cell(row=row_idx, column=6).value = (
                    f'=IF(OR(C{row_idx}="U",C{row_idx}="K",C{row_idx}="D",D{row_idx}=""),'
                    f'{target_hours},'
                    f'IF(C{row_idx}="G",0,(D{row_idx}-C{row_idx})*24-E{row_idx}))'
                )
                ws.cell(row=row_idx, column=7, value=target_hours)
                ws.cell(row=row_idx, column=8).value = (
                    f'=IF(OR(C{row_idx}="U",C{row_idx}="K",C{row_idx}="D"),0,F{row_idx}-G{row_idx})'
                )

                if entry.note:
                    ws.cell(row=row_idx, column=9, value=entry.note)

                for col in range(1, 14):
                    ws.cell(row=row_idx, column=col).border = THIN_BORDER
            else:
                # Leerer Arbeitstag (noch kein Eintrag)
                ws.cell(row=row_idx, column=5, value=default_pause)
                ws.cell(row=row_idx, column=6).value = (
                    f'=IF(OR(C{row_idx}="U",C{row_idx}="K",C{row_idx}="D",D{row_idx}=""),'
                    f'{target_hours},'
                    f'IF(C{row_idx}="G",0,(D{row_idx}-C{row_idx})*24-E{row_idx}))'
                )
                ws.cell(row=row_idx, column=7, value=target_hours)
                ws.cell(row=row_idx, column=8).value = (
                    f'=IF(OR(C{row_idx}="U",C{row_idx}="K",C{row_idx}="D"),0,F{row_idx}-G{row_idx})'
                )
                for col in range(1, 14):
                    ws.cell(row=row_idx, column=col).border = THIN_BORDER

            row_idx += 1

        # Summenzeile
        last_data_row = row_idx - 1
        sum_row = row_idx

        ws.cell(row=sum_row, column=2, value="").font = BOLD_FONT
        ws.cell(row=sum_row, column=5, value="Summen").font = BOLD_FONT
        ws.cell(row=sum_row, column=6).value = f"=SUM(F2:F{last_data_row})"
        ws.cell(row=sum_row, column=6).font = BOLD_FONT
        ws.cell(row=sum_row, column=7).value = f"=SUM(G2:G{last_data_row})"
        ws.cell(row=sum_row, column=7).font = BOLD_FONT
        ws.cell(row=sum_row, column=8).value = f"=SUM(H2:H{last_data_row})"
        ws.cell(row=sum_row, column=8).font = BOLD_FONT

        # Hinweis
        ws.cell(row=sum_row, column=9, value="U=Urlaub, G=Gleittag, K=Krank, D=Dienstreise")

        # Zähler für Abwesenheitstypen
        ws.cell(row=sum_row, column=10).value = f'=COUNTIF(C2:C{last_data_row},"U")'
        ws.cell(row=sum_row, column=11).value = f'=COUNTIF(C2:C{last_data_row},"G")'
        ws.cell(row=sum_row, column=12).value = f'=COUNTIF(C2:C{last_data_row},"K")'
        ws.cell(row=sum_row, column=13).value = f'=COUNTIF(C2:C{last_data_row},"D")'

        # Bedingte Formatierung
        data_range = f"A2:H{last_data_row}"
        ws.conditional_formatting.add(data_range, FormulaRule(formula=['$C2="U"'], fill=TURQUOISE_FILL))
        ws.conditional_formatting.add(data_range, FormulaRule(formula=['$C2="G"'], fill=GLEITTAG_FILL))
        ws.conditional_formatting.add(data_range, FormulaRule(formula=['$C2="K"'], fill=KRANK_FILL))
        ws.conditional_formatting.add(data_range, FormulaRule(formula=['$C2="D"'], fill=DIENSTREISE_FILL))

        # Spaltenbreiten
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 8
        ws.column_dimensions["F"].width = 12
        ws.column_dimensions["G"].width = 8
        ws.column_dimensions["H"].width = 14
        ws.column_dimensions["I"].width = 45

        # Übersicht — Zeile für diesen Monat
        overview_row = month + 1
        uebersicht.cell(row=overview_row, column=1, value=month_name)
        uebersicht.cell(row=overview_row, column=2).value = f"='{month_name}'!G{sum_row}"
        uebersicht.cell(row=overview_row, column=3).value = f"='{month_name}'!F{sum_row}"
        uebersicht.cell(row=overview_row, column=4).value = f"='{month_name}'!H{sum_row}"
        uebersicht.cell(row=overview_row, column=5).value = f"='{month_name}'!J{sum_row}"
        uebersicht.cell(row=overview_row, column=6).value = f"='{month_name}'!L{sum_row}"
        uebersicht.cell(row=overview_row, column=7).value = f"='{month_name}'!M{sum_row}"
        uebersicht.cell(row=overview_row, column=8).value = f"='{month_name}'!K{sum_row}"
        for col in range(1, 9):
            uebersicht.cell(row=overview_row, column=col).border = THIN_BORDER

    # Übersicht — Summenzeile
    sum_overview_row = 14  # Zeile 14 (nach 12 Monaten + Header)
    uebersicht.cell(row=sum_overview_row, column=1, value="Summen").font = BOLD_FONT
    for col in range(2, 9):
        letter = get_column_letter(col)
        uebersicht.cell(row=sum_overview_row, column=col).value = f"=SUM({letter}2:{letter}13)"
        uebersicht.cell(row=sum_overview_row, column=col).font = BOLD_FONT
        uebersicht.cell(row=sum_overview_row, column=col).border = THIN_BORDER

    # Übertrag Vorjahr
    carryover_row = 15
    uebersicht.cell(row=carryover_row, column=1, value="Übertrag Vorjahr").font = BOLD_FONT
    overtime_carryover = get_config_float("overtime_carryover")
    vacation_carryover = get_config_int("vacation_carryover")
    uebersicht.cell(row=carryover_row, column=4, value=overtime_carryover)
    uebersicht.cell(row=carryover_row, column=5, value=vacation_carryover)
    for col in range(1, 9):
        uebersicht.cell(row=carryover_row, column=col).border = THIN_BORDER

    # Gesamtsumme
    total_row = 16
    uebersicht.cell(row=total_row, column=1, value="Gesamtsumme").font = BOLD_FONT
    uebersicht.cell(row=total_row, column=2).value = f"=B{sum_overview_row}"
    uebersicht.cell(row=total_row, column=3).value = f"=C{sum_overview_row}"
    uebersicht.cell(row=total_row, column=4).value = f"=D{sum_overview_row}+D{carryover_row}"
    uebersicht.cell(row=total_row, column=5).value = f"=E{sum_overview_row}-E{carryover_row}"
    uebersicht.cell(row=total_row, column=6).value = f"=F{sum_overview_row}"
    uebersicht.cell(row=total_row, column=7).value = f"=G{sum_overview_row}"
    uebersicht.cell(row=total_row, column=8).value = f"=H{sum_overview_row}"
    for col in range(1, 9):
        uebersicht.cell(row=total_row, column=col).font = BOLD_FONT
        uebersicht.cell(row=total_row, column=col).border = THIN_BORDER

    # Übersicht Spaltenbreiten
    uebersicht.column_dimensions["A"].width = 18
    for col_letter in "BCDEFGH":
        uebersicht.column_dimensions[col_letter].width = 18

    wb.save(output)
    return output
