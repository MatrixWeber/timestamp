"""Import bestehender Excel-Daten (Stundenschreibung) in die SQLite-DB."""

from datetime import date, time, datetime
from pathlib import Path

from openpyxl import load_workbook
from rich.console import Console

from stamp.db import init_db, get_session, Stamp, set_config
from stamp.service import is_workday

console = Console()

MONTHS_EN = [
    "", "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]

TYPE_MAP = {"U": "VACATION", "K": "SICK", "G": "FLEX", "D": "TRAVEL"}


def _parse_excel_time(val) -> time | None:
    """Konvertiert Excel-Zeitwerte zu Python time."""
    if val is None or val == 0 or val == "":
        return None
    if isinstance(val, time):
        return val
    if isinstance(val, datetime):
        return val.time()
    if isinstance(val, str):
        val = val.strip()
        if val in TYPE_MAP or val == "":
            return None
        try:
            return datetime.strptime(val, "%H:%M").time()
        except ValueError:
            try:
                return datetime.strptime(val, "%H:%M:%S").time()
            except ValueError:
                return None
    if isinstance(val, (int, float)):
        if 0 < val < 1:
            total_seconds = int(round(val * 86400))
            hours = total_seconds // 3600
            minutes = (total_seconds % 3600) // 60
            return time(hours, minutes)
    return None


def import_excel(filepath: str, year: int | None = None) -> dict:
    """Importiert Daten aus einer bestehenden Stundenschreibung-Excel."""
    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f"Datei nicht gefunden: {filepath}")

    wb = load_workbook(path, data_only=True)
    init_db()

    if year is None:
        year = date.today().year

    stats = {"imported": 0, "skipped": 0, "errors": 0, "months": 0}

    with get_session() as session:
        for month in range(1, 13):
            month_name = MONTHS_EN[month]
            if month_name not in wb.sheetnames:
                continue

            ws = wb[month_name]
            stats["months"] += 1

            for row_idx in range(2, ws.max_row):
                try:
                    day_val = ws.cell(row=row_idx, column=1).value
                    if day_val is None:
                        continue
                    if isinstance(day_val, datetime):
                        day_num = day_val.day
                    elif isinstance(day_val, (int, float)):
                        day_num = int(day_val)
                    else:
                        continue

                    if day_num < 1 or day_num > 31:
                        continue

                    try:
                        entry_date = date(year, month, day_num)
                    except ValueError:
                        continue

                    # Bestehende Einträge aktualisieren oder neue anlegen
                    existing = session.query(Stamp).filter_by(date=entry_date).first()

                    gekommen = ws.cell(row=row_idx, column=3).value
                    gehzeit = ws.cell(row=row_idx, column=4).value
                    pause_val = ws.cell(row=row_idx, column=5).value
                    hinweis = ws.cell(row=row_idx, column=9).value

                    # Hinweis bereinigen (Summenzeile ignorieren)
                    note = None
                    if hinweis and isinstance(hinweis, str) and "Hinweis:" not in hinweis:
                        note = hinweis.strip()

                    # Check for absence type
                    entry_type = "WORK"
                    if isinstance(gekommen, str) and gekommen.strip() in TYPE_MAP:
                        entry_type = TYPE_MAP[gekommen.strip()]
                        gekommen = None
                        gehzeit = None

                    stamp_in = _parse_excel_time(gekommen)
                    stamp_out = _parse_excel_time(gehzeit)

                    if entry_type == "WORK" and stamp_in is None and stamp_out is None:
                        if not is_workday(entry_date):
                            continue
                        if not existing and not note:
                            stats["skipped"] += 1
                            continue

                    # Pause
                    pause = 0.75
                    if isinstance(pause_val, (int, float)) and pause_val > 0:
                        pause = float(pause_val) if pause_val < 10 else pause_val / 60

                    # Arbeitszeit berechnen
                    work_hours = None
                    overtime = None
                    if entry_type == "WORK" and stamp_in and stamp_out:
                        dt_in = datetime.combine(entry_date, stamp_in)
                        dt_out = datetime.combine(entry_date, stamp_out)
                        work_hours = round((dt_out - dt_in).total_seconds() / 3600 - pause, 2)
                        overtime = round(work_hours - 8.0, 2)
                    elif entry_type != "WORK":
                        work_hours = 8.0
                        overtime = 0.0
                        pause = 0.0

                    if existing:
                        # Update bestehenden Eintrag
                        changed = False
                        if stamp_in and stamp_in != existing.stamp_in:
                            existing.stamp_in = stamp_in
                            changed = True
                        if stamp_out and stamp_out != existing.stamp_out:
                            existing.stamp_out = stamp_out
                            changed = True
                        if entry_type != existing.type:
                            existing.type = entry_type
                            changed = True
                        if note and note != existing.note:
                            existing.note = note
                            changed = True
                        if changed:
                            existing.pause = pause
                            existing.work_hours = work_hours
                            existing.overtime = overtime
                            existing.updated_at = datetime.now()
                            stats["updated"] = stats.get("updated", 0) + 1
                        else:
                            stats["skipped"] += 1
                        continue

                    entry = Stamp(
                        date=entry_date,
                        stamp_in=stamp_in,
                        stamp_out=stamp_out,
                        pause=pause,
                        work_hours=work_hours,
                        overtime=overtime,
                        type=entry_type,
                        note=note,
                    )
                    session.add(entry)
                    stats["imported"] += 1

                except Exception as e:
                    stats["errors"] += 1

            session.commit()

    # Übertrag Vorjahr aus Übersicht lesen
    if "Übersicht" in wb.sheetnames:
        overview = wb["Übersicht"]
        for r in range(overview.max_row, 0, -1):
            label = overview.cell(row=r, column=1).value
            if label and "bertrag" in str(label):
                ot_val = overview.cell(row=r, column=4).value
                if isinstance(ot_val, (int, float)):
                    set_config("overtime_carryover", str(ot_val))
                vac_val = overview.cell(row=r, column=5).value
                if isinstance(vac_val, (int, float)):
                    set_config("vacation_carryover", str(int(vac_val)))
                break

    return stats
